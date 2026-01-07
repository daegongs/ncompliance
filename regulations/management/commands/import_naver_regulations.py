"""
네이버 사규 목록 엑셀 파일에서 데이터를 읽어서 사규 목록을 업데이트하는 명령어
"""

import os
import re
from datetime import datetime, timedelta
from django.core.management.base import BaseCommand
from django.utils import timezone
from django.db import transaction

import openpyxl
from openpyxl.utils import get_column_letter

from accounts.models import Department, User
from regulations.models import Regulation, RegulationVersion


class Command(BaseCommand):
    help = '네이버 사규 목록 엑셀 파일에서 데이터를 읽어서 사규 목록을 업데이트합니다.'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            default='reference/naver_regulation.xlsx',
            help='엑셀 파일 경로 (기본값: reference/naver_regulation.xlsx)'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='실제로 데이터를 저장하지 않고 확인만 합니다.'
        )

    def handle(self, *args, **options):
        file_path = options['file']
        dry_run = options['dry_run']
        
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'파일을 찾을 수 없습니다: {file_path}'))
            return
        
        self.stdout.write(f'엑셀 파일 읽기: {file_path}')
        
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # 헤더 행 확인 (첫 번째 행)
            headers = [cell.value for cell in ws[1]]
            self.stdout.write(f'컬럼 수: {len(headers)}')
            
            # 데이터 읽기
            # 엑셀 컬럼 구조:
            # 0: No, 1: 의무여부, 2: 그룹, 3: 유형, 4: 규정/가이드명
            # 5: 공개여부, 6: 담당부서, 7: 담당자, 8: 제/개정일, 9: 규정본문, 10: 링크
            regulations_data = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0] or str(row[0]).strip() == '':
                    continue
                
                # 데이터 추출
                data = {
                    'no': row[0],
                    'is_mandatory': row[1] if len(row) > 1 else '',  # 의무여부 (의무/비의무)
                    'group': row[2] if len(row) > 2 else '',  # 그룹 (HR, 법무 등)
                    'category': row[3] if len(row) > 3 else '',  # 유형 (규정, 지침 등)
                    'title': row[4] if len(row) > 4 else '',  # 규정/가이드명
                    'is_public': row[5] if len(row) > 5 else '',  # 공개여부
                    'responsible_dept': row[6] if len(row) > 6 else '',  # 담당부서
                    'manager': row[7] if len(row) > 7 else '',  # 담당자
                    'last_revision_date': row[8] if len(row) > 8 else '',  # 제/개정일
                    'content': row[9] if len(row) > 9 else '',  # 규정본문
                    'link': row[10] if len(row) > 10 else '',  # 링크
                }
                
                if not data['title']:
                    continue
                
                regulations_data.append(data)
            
            self.stdout.write(f'총 {len(regulations_data)}개의 사규 데이터를 읽었습니다.')
            
            if dry_run:
                self.stdout.write(self.style.WARNING('DRY-RUN 모드: 데이터를 저장하지 않습니다.'))
                for data in regulations_data[:5]:
                    self.stdout.write(f"  - {data['title']} ({data['category']})")
                return
            
            # 데이터베이스 업데이트
            self.stdout.write('데이터베이스 업데이트 시작...')
            self.import_regulations(regulations_data)
            
            self.stdout.write(self.style.SUCCESS('사규 목록 업데이트 완료!'))
            
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'오류 발생: {e}'))
            import traceback
            traceback.print_exc()

    @transaction.atomic
    def import_regulations(self, regulations_data):
        """사규 데이터를 데이터베이스에 저장합니다."""
        created_count = 0
        updated_count = 0
        error_count = 0
        
        # 기본 사용자 가져오기 (없으면 생성)
        admin_user, _ = User.objects.get_or_create(
            username='admin',
            defaults={
                'email': 'admin@company.com',
                'is_staff': True,
                'is_superuser': True,
            }
        )
        
        # 부서 매핑 딕셔너리 생성
        dept_mapping = self.get_or_create_departments(regulations_data)
        
        for idx, data in enumerate(regulations_data, 1):
            try:
                # 카테고리 매핑 (엑셀 카테고리 → 모델 카테고리)
                category = self.map_category(data['category'])
                
                # 사규 코드 생성 (REG-001 형식)
                code = f"REG-{str(idx).zfill(3)}"
                
                # 책임부서 가져오기
                dept_name = data['responsible_dept'] or 'Compliance'
                responsible_dept = dept_mapping.get(dept_name, dept_mapping.get('Compliance'))
                
                # 날짜 파싱
                effective_date = self.parse_date(data['last_revision_date'])
                
                # 정기검토 예정일 계산 (시행일로부터 1년 후)
                expiry_date = None
                if effective_date:
                    expiry_date = effective_date + timedelta(days=365)
                
                # 상태 결정 (시행일이 있으면 시행중)
                status = 'ACTIVE' if effective_date else 'ACTIVE'
                
                # 의무여부 매핑
                is_mandatory = data['is_mandatory'] == '의무'
                
                # 공개여부 매핑
                is_public = data['is_public'] != '비공개' if data['is_public'] else True
                
                # 사규 생성 또는 업데이트
                regulation, created = Regulation.objects.update_or_create(
                    code=code,
                    defaults={
                        'title': data['title'],
                        'category': category,
                        'is_mandatory': is_mandatory,
                        'scope': 'ALL',
                        'responsible_dept': responsible_dept,
                        'group': data['group'] or '',
                        'manager': data['manager'] or '',
                        'current_version': '1.0',
                        'status': status,
                        'effective_date': effective_date,
                        'expiry_date': expiry_date,
                        'is_public': is_public,
                        'description': '',
                        'reference_url': data['link'] if data['link'] else None,
                        'content': '',  # 본문은 별도로 관리
                        'created_by': admin_user,
                    }
                )
                
                if created:
                    created_count += 1
                    # 초기 버전 생성
                    RegulationVersion.objects.get_or_create(
                        regulation=regulation,
                        version_number='1.0',
                        defaults={
                            'change_type': 'CREATE',
                            'change_reason': '엑셀 데이터에서 가져온 사규',
                            'change_summary': f"최종 개정일: {data['last_revision_date']}",
                            'created_by': admin_user,
                        }
                    )
                else:
                    updated_count += 1
                
                if idx % 50 == 0:
                    self.stdout.write(f'  진행 중... {idx}/{len(regulations_data)}')
                    
            except Exception as e:
                error_count += 1
                self.stdout.write(self.style.WARNING(f'  오류 (행 {idx}): {data.get("title", "N/A")} - {e}'))
        
        self.stdout.write(f'  생성: {created_count}개, 업데이트: {updated_count}개, 오류: {error_count}개')

    def get_or_create_departments(self, regulations_data):
        """부서 매핑 딕셔너리를 생성합니다."""
        dept_mapping = {}
        dept_names = set()
        
        # 모든 부서명 수집
        for data in regulations_data:
            if data['responsible_dept']:
                dept_names.add(data['responsible_dept'])
        
        # 기본 부서들
        default_depts = {
            'Compliance': '준법지원팀',
            'HR': '인사팀',
            'ER Management': '인사팀',
            'IA Plus': '내부감사팀',
            'N FP&A': '재무팀',
            '이사회': '이사회',
            '이사회사무국': '이사회사무국',
        }
        
        # 부서 생성 또는 가져오기
        for dept_name in dept_names:
            if not dept_name:
                continue
            
            # 기본 부서명 매핑
            mapped_name = default_depts.get(dept_name, dept_name)
            
            # 부서 코드 생성
            dept_code = self.generate_dept_code(mapped_name)
            
            dept, _ = Department.objects.get_or_create(
                code=dept_code,
                defaults={
                    'name': mapped_name,
                    'is_active': True,
                }
            )
            dept_mapping[dept_name] = dept
        
        # 기본 Compliance 부서가 없으면 생성
        if 'Compliance' not in dept_mapping:
            compliance_dept, _ = Department.objects.get_or_create(
                code='COMP',
                defaults={
                    'name': '준법지원팀',
                    'is_active': True,
                }
            )
            dept_mapping['Compliance'] = compliance_dept
        
        return dept_mapping

    def generate_dept_code(self, dept_name):
        """부서명에서 부서 코드를 생성합니다."""
        # 한글을 영문 약자로 변환
        mapping = {
            '준법지원팀': 'COMP',
            '인사팀': 'HR',
            '재무팀': 'FIN',
            '내부감사팀': 'IA',
            '이사회': 'BOARD',
            '이사회사무국': 'BOARD_OFF',
        }
        
        if dept_name in mapping:
            return mapping[dept_name]
        
        # 영문 약자 추출
        if len(dept_name) <= 10:
            return dept_name.upper().replace(' ', '_')[:10]
        
        return 'DEPT_' + str(hash(dept_name) % 10000)

    def map_category(self, category_str):
        """엑셀 카테고리를 모델 카테고리로 매핑합니다."""
        # 모델 CATEGORY_CHOICES:
        # ('POLICY', '정책/방침'), ('REGULATION', '규정'),
        # ('GUIDELINE', '지침'), ('MANUAL', '매뉴얼/가이드라인')
        if not category_str:
            return 'REGULATION'
        
        category_str = str(category_str).strip()
        
        mapping = {
            '정책/방침': 'POLICY',
            '규정': 'REGULATION',
            '지침': 'GUIDELINE',
            '매뉴얼/가이드라인': 'MANUAL',
        }
        
        return mapping.get(category_str, 'REGULATION')

    def parse_date(self, date_str):
        """날짜 문자열을 파싱합니다."""
        if not date_str:
            return None
        
        date_str = str(date_str).strip()
        
        # 다양한 날짜 형식 처리
        formats = [
            '%Y. %m. %d.',
            '%Y.%m.%d',
            '%Y-%m-%d',
            '%Y/%m/%d',
            '%Y. %m. %d',
            '%Y.%m.%d.',
        ]
        
        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
        
        # 숫자만 있는 경우 (예: 20250423)
        if date_str.isdigit() and len(date_str) == 8:
            try:
                return datetime.strptime(date_str, '%Y%m%d').date()
            except ValueError:
                pass
        
        # 부분 날짜 처리 (예: 2025.4.14)
        match = re.match(r'(\d{4})\.(\d{1,2})\.(\d{1,2})', date_str)
        if match:
            year, month, day = match.groups()
            try:
                return datetime(int(year), int(month), int(day)).date()
            except ValueError:
                pass
        
        return None
















