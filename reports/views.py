"""
보고서 뷰
사규 현황 및 이력 보고서 생성
"""

import io
from datetime import datetime, timedelta

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Count, Q

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from regulations.models import Regulation, RegulationVersion
from accounts.models import Department


@login_required
def report_index(request):
    """보고서 메인 페이지"""
    return render(request, 'reports/index.html')


@login_required
def regulation_status_report(request):
    """사규 현황 보고서"""
    # 필터 파라미터
    category = request.GET.get('category', '')
    status = request.GET.get('status', '')
    dept_id = request.GET.get('department', '')
    
    regulations = Regulation.objects.select_related('responsible_dept').all()
    
    if category:
        regulations = regulations.filter(category=category)
    if status:
        regulations = regulations.filter(status=status)
    if dept_id:
        regulations = regulations.filter(responsible_dept_id=dept_id)
    
    # 통계
    total_count = regulations.count()
    category_stats = regulations.values('category').annotate(count=Count('id'))
    status_stats = regulations.values('status').annotate(count=Count('id'))
    
    context = {
        'regulations': regulations.order_by('category', 'code'),
        'total_count': total_count,
        'category_stats': category_stats,
        'status_stats': status_stats,
        'departments': Department.objects.filter(is_active=True),
        'selected_category': category,
        'selected_status': status,
        'selected_dept': dept_id,
    }
    
    return render(request, 'reports/regulation_status.html', context)


@login_required
def export_regulation_status_excel(request):
    """사규 현황 보고서 엑셀 다운로드"""
    # 필터 파라미터
    category = request.GET.get('category', '')
    status = request.GET.get('status', '')
    dept_id = request.GET.get('department', '')
    
    regulations = Regulation.objects.select_related('responsible_dept').all()
    
    if category:
        regulations = regulations.filter(category=category)
    if status:
        regulations = regulations.filter(status=status)
    if dept_id:
        regulations = regulations.filter(responsible_dept_id=dept_id)
    
    regulations = regulations.order_by('category', 'code')
    
    # Excel 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "사규 현황"
    
    # 스타일 정의
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 헤더
    headers = ['No.', '사규코드', '사규명', '분류', '상태', '의무준수', '적용범위', '책임부서', '현재버전', '시행일', '정기검토예정일']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 분류 레이블
    category_labels = dict(Regulation.CATEGORY_CHOICES)
    status_labels = dict(Regulation.STATUS_CHOICES)
    scope_labels = dict(Regulation.SCOPE_CHOICES)
    
    # 데이터
    for row, reg in enumerate(regulations, 2):
        ws.cell(row=row, column=1, value=row-1).border = thin_border
        ws.cell(row=row, column=2, value=reg.code).border = thin_border
        ws.cell(row=row, column=3, value=reg.title).border = thin_border
        ws.cell(row=row, column=4, value=category_labels.get(reg.category, reg.category)).border = thin_border
        ws.cell(row=row, column=5, value=status_labels.get(reg.status, reg.status)).border = thin_border
        ws.cell(row=row, column=6, value='의무' if reg.is_mandatory else '비의무').border = thin_border
        ws.cell(row=row, column=7, value=scope_labels.get(reg.scope, reg.scope)).border = thin_border
        ws.cell(row=row, column=8, value=reg.responsible_dept.name if reg.responsible_dept else '').border = thin_border
        ws.cell(row=row, column=9, value=reg.current_version).border = thin_border
        ws.cell(row=row, column=10, value=reg.effective_date.strftime('%Y-%m-%d') if reg.effective_date else '').border = thin_border
        ws.cell(row=row, column=11, value=reg.expiry_date.strftime('%Y-%m-%d') if reg.expiry_date else '').border = thin_border
    
    # 컬럼 너비 조정
    column_widths = [6, 15, 40, 15, 10, 10, 12, 20, 10, 12, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # 응답 생성
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"사규현황보고서_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
def change_history_report(request):
    """제개정 이력 보고서"""
    user = request.user
    # 기간 필터
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    change_type = request.GET.get('change_type', '')
    
    versions = RegulationVersion.objects.select_related(
        'regulation', 'regulation__responsible_dept', 'created_by', 'approved_by'
    ).filter(regulation__isnull=False)
    
    # 책임부서담당자(DEPT_MANAGER)인 경우: 자신이 담당하는 사규만 표시
    if user.role == 'DEPT_MANAGER' and user.department:
        versions = versions.filter(
            Q(regulation__responsible_dept=user.department) |
            Q(regulation__manager__icontains=user.get_full_name()) |
            Q(regulation__manager_primary__icontains=user.get_full_name())
        )
    
    if start_date:
        versions = versions.filter(created_at__date__gte=start_date)
    if end_date:
        versions = versions.filter(created_at__date__lte=end_date)
    if change_type:
        versions = versions.filter(change_type=change_type)
    
    versions = versions.order_by('-created_at')
    
    context = {
        'versions': versions,
        'total_count': versions.count(),
        'selected_start': start_date,
        'selected_end': end_date,
        'selected_type': change_type,
        'is_dept_manager': user.role == 'DEPT_MANAGER',
    }
    
    return render(request, 'reports/change_history.html', context)


@login_required
def export_change_history_excel(request):
    """제개정 이력 보고서 엑셀 다운로드"""
    user = request.user
    # 기간 필터
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    change_type = request.GET.get('change_type', '')
    
    versions = RegulationVersion.objects.select_related(
        'regulation', 'regulation__responsible_dept', 'created_by', 'approved_by'
    ).filter(regulation__isnull=False)
    
    # 책임부서담당자(DEPT_MANAGER)인 경우: 자신이 담당하는 사규만
    if user.role == 'DEPT_MANAGER' and user.department:
        versions = versions.filter(
            Q(regulation__responsible_dept=user.department) |
            Q(regulation__manager__icontains=user.get_full_name()) |
            Q(regulation__manager_primary__icontains=user.get_full_name())
        )
    
    if start_date:
        versions = versions.filter(created_at__date__gte=start_date)
    if end_date:
        versions = versions.filter(created_at__date__lte=end_date)
    if change_type:
        versions = versions.filter(change_type=change_type)
    
    versions = versions.order_by('-created_at')
    
    # Excel 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "제개정 이력"
    
    # 스타일
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 헤더
    headers = ['No.', '사규코드', '사규명', '버전', '변경유형', '변경사유', '작성자', '승인자', '승인일', '등록일']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    change_type_labels = dict(RegulationVersion.CHANGE_TYPE_CHOICES)
    
    # 데이터
    for row, ver in enumerate(versions, 2):
        ws.cell(row=row, column=1, value=row-1).border = thin_border
        ws.cell(row=row, column=2, value=ver.regulation.code if ver.regulation else '삭제된 사규').border = thin_border
        ws.cell(row=row, column=3, value=ver.regulation.title if ver.regulation else '-').border = thin_border
        ws.cell(row=row, column=4, value=f"v{ver.version_number}").border = thin_border
        ws.cell(row=row, column=5, value=change_type_labels.get(ver.change_type, ver.change_type)).border = thin_border
        ws.cell(row=row, column=6, value=ver.change_reason[:50] + '...' if len(ver.change_reason) > 50 else ver.change_reason).border = thin_border
        ws.cell(row=row, column=7, value=ver.created_by.get_full_name() if ver.created_by else '').border = thin_border
        ws.cell(row=row, column=8, value=ver.approved_by.get_full_name() if ver.approved_by else '').border = thin_border
        ws.cell(row=row, column=9, value=ver.approved_at.strftime('%Y-%m-%d') if ver.approved_at else '').border = thin_border
        ws.cell(row=row, column=10, value=ver.created_at.strftime('%Y-%m-%d %H:%M')).border = thin_border
    
    # 컬럼 너비 조정
    column_widths = [6, 15, 40, 10, 10, 50, 15, 15, 12, 18]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # 응답 생성
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"제개정이력보고서_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
def department_report(request):
    """부서별 보고서"""
    dept_id = request.GET.get('department', '')
    
    departments = Department.objects.filter(is_active=True).annotate(
        regulation_count=Count('regulations')
    ).order_by('-regulation_count')
    
    regulations = None
    selected_dept = None
    
    if dept_id:
        selected_dept = Department.objects.filter(pk=dept_id).first()
        if selected_dept:
            regulations = Regulation.objects.filter(
                responsible_dept=selected_dept
            ).order_by('category', 'code')
    
    context = {
        'departments': departments,
        'regulations': regulations,
        'selected_dept': selected_dept,
    }
    
    return render(request, 'reports/department_report.html', context)


@login_required
def expiry_report(request):
    """만료예정 보고서"""
    user = request.user
    days = int(request.GET.get('days', 30))
    
    today = timezone.now().date()
    target_date = today + timedelta(days=days)
    
    regulations = Regulation.objects.filter(
        status='ACTIVE',
        expiry_date__lte=target_date,
        expiry_date__gte=today
    ).select_related('responsible_dept')
    
    # 책임부서담당자(DEPT_MANAGER)인 경우: 자신이 담당하는 사규만 표시
    if user.role == 'DEPT_MANAGER' and user.department:
        regulations = regulations.filter(
            Q(responsible_dept=user.department) |
            Q(manager__icontains=user.get_full_name()) |
            Q(manager_primary__icontains=user.get_full_name())
        )
    
    regulations = regulations.order_by('expiry_date')
    
    context = {
        'regulations': regulations,
        'total_count': regulations.count(),
        'selected_days': days,
        'target_date': target_date,
        'is_dept_manager': user.role == 'DEPT_MANAGER',
    }
    
    return render(request, 'reports/expiry_report.html', context)


@login_required
def export_expiry_excel(request):
    """만료예정 보고서 엑셀 다운로드"""
    user = request.user
    days = int(request.GET.get('days', 30))
    
    today = timezone.now().date()
    target_date = today + timedelta(days=days)
    
    regulations = Regulation.objects.filter(
        status='ACTIVE',
        expiry_date__lte=target_date,
        expiry_date__gte=today
    ).select_related('responsible_dept')
    
    # 책임부서담당자(DEPT_MANAGER)인 경우: 자신이 담당하는 사규만
    if user.role == 'DEPT_MANAGER' and user.department:
        regulations = regulations.filter(
            Q(responsible_dept=user.department) |
            Q(manager__icontains=user.get_full_name()) |
            Q(manager_primary__icontains=user.get_full_name())
        )
    
    regulations = regulations.order_by('expiry_date')
    
    # Excel 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "만료예정 사규"
    
    # 스타일
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 헤더
    headers = ['No.', '사규코드', '사규명', '분류', '책임부서', '정기검토예정일', '남은일수']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    category_labels = dict(Regulation.CATEGORY_CHOICES)
    
    # 데이터
    for row, reg in enumerate(regulations, 2):
        days_left = (reg.expiry_date - today).days
        ws.cell(row=row, column=1, value=row-1).border = thin_border
        ws.cell(row=row, column=2, value=reg.code).border = thin_border
        ws.cell(row=row, column=3, value=reg.title).border = thin_border
        ws.cell(row=row, column=4, value=category_labels.get(reg.category, reg.category)).border = thin_border
        ws.cell(row=row, column=5, value=reg.responsible_dept.name if reg.responsible_dept else '').border = thin_border
        ws.cell(row=row, column=6, value=reg.expiry_date.strftime('%Y-%m-%d')).border = thin_border
        ws.cell(row=row, column=7, value=f"{days_left}일").border = thin_border
    
    # 컬럼 너비 조정
    column_widths = [6, 15, 40, 15, 20, 15, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # 응답 생성
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"만료예정보고서_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response






